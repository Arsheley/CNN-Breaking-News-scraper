import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os


def fetch_news():
    url = "https://search.prod.di.api.cnn.io/content?q=breaking%20news%20&size=10&from=0&page=1&sort=newest&request_id=stellar-search-e1a7628f-9e56-446c-a032-0783b2a78c26&site=cnn&zaid=e69ff0bf-726a-44c1-b4ec-24403ae7d6e2"
    headers = {
    "User-Agent": "Mozilla/5.0"
}

    try:
        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()
        data = response.json()
        return data.get("result", [])

    except Exception as e:
        print(f"Error fetching data: {e}")
        return []


def load_or_create_workbook(filename):
    if os.path.exists(filename):
        wb = load_workbook(filename)
    else:
        wb = Workbook()
        sheet = wb.active
        sheet.title = "CNN Breaking News"

        sheet.append(["Headline", "Last Modified Date", "Body", "Scraped Time"])

    return wb


def get_existing_headlines(sheet):
    headlines = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            headlines.add(row[0])
    return headlines


def format_sheet(sheet):
    # Alignment
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                wrap_text=True,
                vertical="top",
                horizontal="left"
            )

    # Column width adjustment
    for column in sheet.columns:
        max_length = 0
        col_letter = get_column_letter(column[0].column)

        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        sheet.column_dimensions[col_letter].width = min(max_length + 2, 50)

    # Header styling
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    # Borders for all cells
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border

    sheet.freeze_panes = "A2"


def save_news():
    filename = "CNN_Breaking_News.xlsx"
    wb = load_or_create_workbook(filename)
    sheet = wb.active

    existing_headlines = get_existing_headlines(sheet)
    articles = fetch_news()
    print(len(articles))
    scraped_time = datetime.now()

    added_count = 0

    for article in articles:
        headline = article.get("headline")

        if not headline:
            continue
        print(headline)
        if headline not in existing_headlines:
           
            sheet.append([
                headline,
                article.get("lastModifiedDate", ""),
                article.get("body", ""),
                scraped_time
            ])

            existing_headlines.add(headline)
            added_count += 1

    format_sheet(sheet)

    wb.save(filename) 

    print(f"Done! Added {added_count} new articles.")


# Run the program
save_news()
 

